#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import datetime
import csv
import shutil
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from config import config

from download.box import LifespanBox


# In[2]:


config['root'] = {'cache': '/home/osboxes/PycharmProjects/ccf/tmp/cache/', 'store': '/home/osboxes/PycharmProjects/ccf/tmp/store/'}


# In[7]:


box_config = '/home/osboxes/PycharmProjects/ccf/tmp/.boxApp/config.json'


# In[4]:


from boxsdk import JWTAuth, OAuth2, Client


# In[5]:


JWTAuth


# In[8]:


verbose = True
snapshotdate = datetime.datetime.today().strftime('%m_%d_%Y')
root_cache = config['root']['cache']
ksads_cache_path = os.path.join(root_cache, 'ksads')
if not os.path.exists(ksads_cache_path):
    os.mkdir(ksads_cache_path)

root_store = config['root']['store']
# this will be the place to save any snapshots on the nrg servers
store_space = os.path.join(root_store, 'ksads')
if not os.path.exists(store_space):
    os.mkdir(store_space)

# connect to Box
box = LifespanBox(cache=ksads_cache_path, config_file=box_config)

# snapshot folder (used to be the combined folder)
ksads_snapshotfolderid = 48203202724
snapshotQCfolder = 76434619813


# In[9]:


keyfileid = box.search(pattern='*Key')


# In[1]:


import os
import datetime
import csv
import shutil
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from config import config

from download.box import LifespanBox



config['root'] = {'cache': '/home/osboxes/PycharmProjects/ccf/tmp/cache/',
                  'store': '/home/osboxes/PycharmProjects/ccf/tmp/store/'}
config['config_files']['box'] = '/home/osboxes/PycharmProjects/ccf/tmp/.boxApp/config.json'

# verbose = False
verbose = True
snapshotdate = datetime.datetime.today().strftime('%m_%d_%Y')
root_cache = config['root']['cache']
ksads_cache_path = os.path.join(root_cache, 'ksads')
if not os.path.exists(ksads_cache_path):
    os.mkdir(ksads_cache_path)

root_store = config['root']['store']
# this will be the place to save any snapshots on the nrg servers
store_space = os.path.join(root_store, 'ksads')
if not os.path.exists(store_space):
    os.mkdir(store_space)

# connect to Box
box = LifespanBox(cache=ksads_cache_path, config_file=config['config_files']['box'])

# snapshot folder (used to be the combined folder)
ksads_snapshotfolderid = 48203202724
snapshotQCfolder = 76434619813

# download one of the identical key files which contain the labels for
# all the numbered questions in KSADS
cachekeyfile = box.downloadFile(506958838440)

# hard coding to prevent read of files that shouldnt be in these folders in box.  cant do a search.
# WU,UMN,UCLA, and Harvard, respectively, for cleanest start file ids below
assessments = config['Assessments']


# In[3]:


for item in assessments:
    print(item)


# In[4]:


item = 'Intro'


# In[5]:


# Download latest clean (box) files for each site - make sure they exist.  send up warning if they dont -
# Before running this program, Montly update process is to download data (3 files - intro,screener,supplement,
# for each site) from KSADS.net, append rows to site/assessment file (by hand until this program written to automate),
# and incorporate and errors from wiki...look for any missing rows, but check redcap visit summary to make sure not already known to be missing
# then save updated xls files with new date, navigate to file in box and select 'upload new version.'
# This will allow BOX to track versioning until better system is in place
site_files = assessments[item]['cleanest_start']

# Get all rows from all site output files for cleanest files as a
# pandas dataframe with column labels
rows = get_all_rows(site_files)


# In[6]:


def get_all_rows(sites):
    """
    reads all rows from site xls files as data frame
    """
    rows = []
    for site_file in sites:
        # Download file contents to cache
        path = box.downloadFile(site_file)
        wb = load_workbook(filename=path)
        # print(wb.sheetnames)
        if len(wb.sheetnames) > 1:
            print('More than one worksheet.')
            print(wb.sheetnames)
            print('Using the first one --> ' + wb.sheetnames[0])
            continue
        questionnaire = wb[wb.sheetnames[0]]
        print(questionnaire)
        # Skip the header for all but the first file
        current_row = 0
        for row in questionnaire.values:
            if current_row == 0:
                header = row
            if current_row != 0:
                rows.append(row)
            current_row += 1
        rowsasdf = pd.DataFrame(rows, columns=header)
    return rowsasdf


# In[29]:


def get_all_rows(sites):
    """
    reads all rows from site xls files as data frame
    """
    dataframes = []
    for site_file in sites:
        # Download file contents to cache
        path = box.downloadFile(site_file)
        df = pd.read_excel(path)
        df.columns = df.columns.astype(str)
        dataframes.append(df)
        
    return pd.concat(dataframes)


# In[39]:


snap = 'KSADS_' + item + '_Snapshot_' + snapshotdate + '.csv'
snapshot_filepath = os.path.join(store_space, snap)  # or in store_space?
QC_filepath = os.path.join(ksads_cache_path, 'QC_' + snap)
dictcsv_filepath = os.path.join(ksads_cache_path, 'Dict_' + snap)

# write rows to csv in store
rows.to_csv(snapshot_filepath, index=False)
# upload the snapshot into box
box.upload_file(snapshot_filepath, ksads_snapshotfolderid)


# In[ ]:


# compare ids from snapshot (currently loaded into 'rows' dataframe)
# with those in Redcap.
studyids = box.getredcapids()
studydata = box.getredcapdata()
rowsofinterest = rows[['ID', 'PatientID',
                       'PatientType', 'SiteName']].copy()
new = rowsofinterest['PatientID'].str.split("_", 1, expand=True)
rowsofinterest['subject'] = new[0].str.strip()
combowredcap = pd.merge(
    rowsofinterest,
    studyids,
    how='left',
    on='subject')
# these are the ids that need to be checked by the sites
combonotinredcap = combowredcap.loc[combowredcap.Subject_ID.isnull()].copy()
combonotinredcap['reason'] = 'PatientID not in Redcap'
# combonotinredcap[['ID','PatientID','PatientType','SiteName','reason']].to_csv(QC_filepath,index=False)


# In[11]:


path = '../tmp/cache/ksads/Harvard_HCPD KSADS Intro_2019Oct28.xlsx'


# In[12]:


pd.read_excel(path)


# In[30]:


r = get_all_rows(site_files)


# In[31]:


r.shape


# In[21]:


r.columns


# In[35]:


r.set_index('PatientID').sort_index() == rows.set_index('PatientID').sort_index()


# In[37]:


r[r.PatientID == 'HCD2996590_V1']


# In[38]:


rows[rows.PatientID == 'HCD2996590_V1']


# In[ ]:


row


# In[14]:


rows.shape


# In[3]:


import pandas as pd
import pycurl
import json
from io import BytesIO
import numpy as np
from config import config

redcapconfigfile = '../tmp/.boxApp/redcapconfig.csv'
redcap_api_url = config['redcap']['api_url']


# In[4]:


"""
Downloads field (IDS) in Redcap databases specified by details in redcapconfig file
Returns panda dataframe with fields 'study', 'Subject_ID, 'subject', and 'flagged', where 'Subject_ID' is the
patient id in the database of interest (sometimes called subject_id, parent_id).
subject is this same id stripped of underscores or flags like 'excluded' to make it easier to merge
flagged contains the extra characters other than the id so you can keep track of who should NOT be uploaded to NDA
 or elsewwhere shared
"""
auth = pd.read_csv(redcapconfigfile)
studyids = pd.DataFrame()


# In[8]:


for z in auth.to_dict(orient='records'):
    pass


# In[45]:


auth


# In[55]:


auth.set_index('study',drop=False).to_dict('index')


# In[44]:


auth.columns


# In[40]:


token


# In[16]:


data = {
    'token': z['token'],
    'content': 'record',
    'format': 'csv',
    'type': 'flat',
    'fields': [z['field']],
    'events': [z['event']],
    'rawOrLabel': 'raw',
    'rawOrLabelHeaders': 'raw',
    'exportCheckboxLabel': 'false',
    'exportSurveyFields': 'false',
    'exportDataAccessGroups': 'false',
    'returnFormat': 'json'}


# In[11]:


import requests


# In[26]:


r = requests.post(redcap_api_url, data=data)


# In[20]:


r = r.content.decode('UTF-8')


# In[29]:


import io


# In[31]:


df = pd.read_csv(io.BytesIO(r.content), encoding='utf8')


# In[32]:


df


# In[ ]:


parent_ids = pd.DataFrame(
        htmlString.splitlines(),
        columns=['Subject_ID'])
    parent_ids = parent_ids.iloc[1:, ]
    parent_ids = parent_ids.loc[~(parent_ids.Subject_ID == '')]
    uniqueids = pd.DataFrame(
        parent_ids.Subject_ID.unique(),
        columns=['Subject_ID'])
    uniqueids['Subject_ID'] = uniqueids.Subject_ID.str.strip('\'"')
    new = uniqueids['Subject_ID'].str.split("_", 1, expand=True)
    uniqueids['subject'] = new[0].str.strip()
    uniqueids['flagged'] = new[1].str.strip()
    uniqueids['study'] = auth.study[i]
    studyids = studyids.append(uniqueids)


# In[28]:


for i in range(len(auth.study)):
    data = {
        'token': auth.token[i],
        'content': 'record',
        'format': 'csv',
        'type': 'flat',
        'fields[0]': auth.field[i],
        'events[0]': auth.event[i],
        'rawOrLabel': 'raw',
        'rawOrLabelHeaders': 'raw',
        'exportCheckboxLabel': 'false',
        'exportSurveyFields': 'false',
        'exportDataAccessGroups': 'false',
        'returnFormat': 'json'}
    buf = BytesIO()
    ch = pycurl.Curl()
    ch.setopt(
        ch.URL,
        redcap_api_url)
    ch.setopt(ch.HTTPPOST, list(data.items()))
    ch.setopt(ch.WRITEDATA, buf)
    ch.perform()
    ch.close()
    htmlString = buf.getvalue().decode('UTF-8')
    buf.close()
    parent_ids = pd.DataFrame(
        htmlString.splitlines(),
        columns=['Subject_ID'])
    parent_ids = parent_ids.iloc[1:, ]
    parent_ids = parent_ids.loc[~(parent_ids.Subject_ID == '')]
    uniqueids = pd.DataFrame(
        parent_ids.Subject_ID.unique(),
        columns=['Subject_ID'])
    uniqueids['Subject_ID'] = uniqueids.Subject_ID.str.strip('\'"')
    new = uniqueids['Subject_ID'].str.split("_", 1, expand=True)
    uniqueids['subject'] = new[0].str.strip()
    uniqueids['flagged'] = new[1].str.strip()
    uniqueids['study'] = auth.study[i]
    studyids = studyids.append(uniqueids)
return studyids


# In[ ]:




