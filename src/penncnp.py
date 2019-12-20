import os
import datetime
import csv
import pycurl
import sys
import shutil
from openpyxl import load_workbook
import pandas as pd
import numpy as np

from download import redcap
from download.redcap import Redcap

redcap = Redcap('../tmp/.boxApp/redcapconfig.csv')
from download.box import LifespanBox


verbose = True
#verbose = False
snapshotdate = datetime.datetime.today().strftime('%m_%d_%Y')

root_cache = '/nrgpackages/tools.release/intradb/ccf-nda-behavioral/cache/'
cache_space = os.path.join(root_cache, 'penncnp')
try:
    os.mkdir(cache_space)
except BaseException:
    print("cache already exists")

root_store = '/home/shared/HCP/hcpinternal/ccf-nda-behavioral/store/'
# this will be the place to save any snapshots on the nrg servers
store_space = os.path.join(root_store, 'penncnp')
try:
    os.mkdir(store_space)  # look for store space before creating it here
except BaseException:
    print("store already exists")

# connect to Box
box = LifespanBox(cache=cache_space)

# snapshot folder (used to be the combined folder)
penn_snapshotfolderid = 48203214504
snapshotQCfolder = 76434619813
slimfolder = 77127935742


# hard coding to prevent read of files that shouldnt be in these folders in box.  cant do a search.
# WU,UMN,UCLA, and Harvard, HCA and HCD data all in one single file under
# Behavioral Data - allsites . This is the 'database' for PennCNP
assessments = {
    'PennCNP': {
        'slim_id': 460781218444,
        'dict_id': 460884641280,
        'cleanest_start': [452784840845]
    }
}


def main():
    for item in assessments:
        # Download latest clean (box) allsites file - make sure they exist.  send up warning if it dont -
        # Coordinator monthly update process is to download data (1 file contains all site data -convenient) from penncnp.med.upenn.edu/results.pl ,
        # append any new rows to site/assessment file (by downloading box file), incorporate and errors from wiki...and/or hca data checklist
        # then save updated xls files with new date, navigate to file in box and select 'upload new version.'
        # This will allow BOX to track versioning until better system is in place
        # KIDS UNDER 8 dont do PENNCNP
        site_files = assessments[item]['cleanest_start']
        # Get all rows from all site output files for cleanest files as a
        # pandas dataframe with column labels
        rows = get_all_rows(site_files)
        # create snapshot of combined file store snapshot in 'store' and in
        # 'snapshots' under all sites directory in box.
        snap = 'PennCNP_' + item + '_Snapshot_' + snapshotdate + '.csv'
        snapshotfile = os.path.join(store_space, snap)
        QCfile = os.path.join(cache_space, 'QC_' + snap)
        # write rows to csv in store
        rows.to_csv(snapshotfile, index=False)
        # upload the snapshot into box
        box.upload_file(snapshotfile, penn_snapshotfolderid)
        # compare ids from snapshot (currently loaded into 'rows' dataframe)
        # with those in Redcap.
        studyids = redcap.getredcapids()
        studydata = redcap.getredcapdata()
        # rowsofinterest=rows[['ID','PatientID','PatientType','SiteName']].copy()
        rowsofinterest = rows[['datasetid',
                               'siteid', 'subid', 'assessment']].copy()
        rowsofinterest['subject'] = rowsofinterest.subid.str.strip()
        combowredcap = pd.merge(
            rowsofinterest,
            studyids,
            how='left',
            on='subject')
        # these are the ids that need to be checked by the sites
        combonotinredcap = combowredcap.loc[combowredcap.Subject_ID.isnull()].copy(
        )
        combonotinredcap['reason'] = 'PatientID not in Redcap'
        # combonotinredcap[['datasetid','subject','assessment','siteid','reason']].to_csv(QCfile,index=False)
        combowredcap2 = pd.merge(
            rowsofinterest,
            studydata,
            how='right',
            on='subject')
        combonotinbox = combowredcap2.loc[combowredcap2.subid.isnull()]
        notinboxunique = combonotinbox.drop_duplicates('subject_id')
        #notinboxunique.loc[(notinboxunique.interview_date<'2019-05-01') & (notinboxunique.flagged.isnull()==True)]
        notinboxunique = notinboxunique.loc[notinboxunique.flagged.isnull()]
        # notinboxunique=notinboxunique.loc[notinboxunique.interview_date<'2019-05-01']
        status1 = redcap.getredcapfields(
            ['data_status', 'misscat'], study='hcpdchild')
        status1 = status1[['data_status', 'misscat___9', 'subject_id']].copy()
        tnotinboxunique = pd.merge(
            notinboxunique,
            status1,
            how='left',
            on='subject_id')
        status2 = redcap.getredcapfields(['data_status', 'misscat'], study='hcpa')
        status2 = status2[['data_status', 'misscat___7', 'subject_id']].copy()
        tnotinboxunique = pd.merge(
            tnotinboxunique,
            status2,
            how='left',
            on='subject_id')
        tnotinboxunique['misscat'] = tnotinboxunique['misscat___9']
        tnotinboxunique.loc[tnotinboxunique.misscat.isnull(),
                            'misscat'] = tnotinboxunique['misscat___7']
        tnotinboxunique.loc[tnotinboxunique.data_status_x.isnull(
        ), 'data_status_x'] = tnotinboxunique.data_status_y
        status3 = redcap.getredcapfields(
            ['data_status', 'misscat'], study='hcpd18')
        status3 = status3[['data_status', 'subject_id', 'misscat___9']].copy()
        tnotinboxunique = pd.merge(
            tnotinboxunique.drop(
                columns={
                    'data_status_y',
                    'misscat___7',
                    'misscat___9'}),
            status3,
            how='left',
            on='subject_id')
        tnotinboxunique.loc[tnotinboxunique.data_status_x.isnull(
        ), 'data_status_x'] = tnotinboxunique.data_status
        tnotinboxunique.loc[tnotinboxunique.misscat.isnull(),
                            'misscat'] = tnotinboxunique['misscat___9']
        notinboxunique = tnotinboxunique.drop(
            columns={
                'data_status',
                'misscat___9'}).rename(
            columns={
                'data_status_x': 'data_status'})
        # this code needs to be moved to box.py next iteration - exclude
        # subjects who are too young to have had the test
        datevar = 'interview_date'
        notinboxunique['nb_months'] = (12 *
                                       (pd.to_datetime(notinboxunique[datevar]).dt.year -
                                        pd.to_datetime(notinboxunique.dob).dt.year) +
                                       (pd.to_datetime(notinboxunique[datevar]).dt.month -
                                           pd.to_datetime(notinboxunique.dob).dt.month) +
                                       (pd.to_datetime(notinboxunique[datevar]).dt.day -
                                           pd.to_datetime(notinboxunique.dob).dt.day) /
                                       31)
        notinboxunique['nb_months'] = notinboxunique['nb_months'].apply(
            np.floor).astype(int)
        notinboxunique = notinboxunique.loc[notinboxunique.nb_months >= 96]
        notinboxunique.loc[notinboxunique.subject_id != 'HCDJanePractice']

        # drop ids known to be permanently missing
        perm_misslist = [
            'HCA6012744',
            'HCA6605569',
            'HCA8689410',
            'HCA8735289',
            'HCA8917700',
            'HCA9695713',
            'HCD1539759',
            'HCD1714852',
            'HCD1769978',
            'HCD2642858',
            'HCD2913459',
            'HCD2978386']
        notinboxunique = notinboxunique.loc[~(
            notinboxunique.subject.isin(perm_misslist))]
        # subset to ids listed as data_status=2, misscat=0 - these are the
        # folks that should be
        notinboxunique = notinboxunique.loc[notinboxunique.interview_date < '2019-08-13']
        notinboxunique = notinboxunique.loc[(
            notinboxunique.data_status == 2) & (notinboxunique.misscat == 0)]

        notinboxunique['reason'] = 'Missing in Box'
        #####################
        # now find duplicates and missing visits
        dupls = rows.loc[rows.duplicated(subset=['subid', 'assessment'], keep=False)][[
            'datasetid', 'siteid', 'subid', 'assessment', 'sex', 'age', 'handedness']]
        dupls['reason'] = 'duplicated subid/assessment combination'
        checkvisit = rows.loc[rows.assessment.isin(['V1', 'V2', 'V3']) == False][[
            'datasetid', 'siteid', 'subid', 'assessment', 'sex', 'age', 'handedness']]
        checkvisit['reason'] = 'visit number not in V1, V2, or V3'

        catQC = pd.concat([combonotinredcap,
                           notinboxunique,
                           dupls,
                           checkvisit],
                          axis=0,
                          sort=True)[['datasetid',
                                      'siteid',
                                      'subid',
                                      'assessment',
                                      'subject',
                                      'study',
                                      'site',
                                      'gender',
                                      'interview_date',
                                      'reason']]
        catQC = catQC.sort_values(['site', 'study'])
        catQC.to_csv(QCfile, index=False)

        # upload QC file to box
        box.upload_file(QCfile, snapshotQCfolder)
        # just keep these in mind for time being
        # these are the ids that need to be excluded before any data is
        # shared...make note.
        comboflaggedinredcap = combowredcap.loc[combowredcap.flagged.isnull(
        ) == False]
        # NDA release will be based on an explicitly named snapshot...not just whatever is in the rows right now - but the following files will
        # ease the process of harmonization.
        # this is a temporary file in the Behavioral Data - Snapshots / PENN /
        # temp folder for use viewing data
        slimhandle = assessments[item]['slim_id']
        # make slim file in cache from snapshot in store (which is identical
        # copy with same named file on box).
        slimf = makeslim(snapshotfile, slimhandle)
        # Note this will overwrite whatever is in the slim file with current data - write slim file to cache and to Scratch Export to NDA
        # make a draft datadictionary from slim file and start comparing columns - write dictionary to cache and to Scratch Export to NDA
        # makedatadict(slimf,dict_id=assessments[item]['dict_id'],folderout=slimfolder)#,assessments[item]['dict_id'])#,cachekeyfile,assessments[item]['key_sheet'])
    # Clean up cache space
    shutil.rmtree(box.cache)




def get_all_rows(sites):
    """
    reads all rows from site xls files as data frame
    """
    rows = []
    for site_file in sites:
        # Download file contents to cache
        path = box.downloadFile(site_file)
        wb = load_workbook(filename=path)
        # print(wb.sheetnames)
        if len(wb.sheetnames) > 1:
            print('More than one worksheet.')
            print(wb.sheetnames)
            print('Using the first one --> ' + wb.sheetnames[0])
            continue
        questionnaire = wb[wb.sheetnames[0]]
        print(questionnaire)
        # Skip the header for all but the first file
        current_row = 0
        for row in questionnaire.values:
            if current_row == 0:
                header = row
            if current_row != 0:
                rows.append(row)
            current_row += 1
        rowsasdf = pd.DataFrame(rows, columns=header)
    return rowsasdf



def makeslim(storefilename, slim_id):
    """
    remove columns from cachecopy that have no data and upload slim file to box
    """
    box.downloadFile(slim_id)
    slimf = box.getFileById(slim_id)
    ksadsraw = pd.read_csv(storefilename, header=0, low_memory=False)
    ksadsraw = ksadsraw.dropna(axis=1, how='all')
    snapname = os.path.basename(storefilename)
    combined_fileout = os.path.join(
        box.cache, snapname.split('.')[0] + 'Slim.csv')
    ksadsraw.to_csv(combined_fileout, index=False)
    # box.client.folder(str(ksadscombinedfolderid)).upload(fileout)
    slimf.update_contents(combined_fileout)
    return slimf




if __name__ == '__main__':
    main()
