import os
import datetime
import csv
import shutil
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from config import config
from download import redcap

from download.box import LifespanBox



config['root'] = {'cache': '/home/osboxes/PycharmProjects/ccf/tmp/cache/',
                  'store': '/home/osboxes/PycharmProjects/ccf/tmp/store/'}
config['config_files']['box'] = '/home/osboxes/PycharmProjects/ccf/tmp/.boxApp/config.json'

# verbose = False
verbose = True
snapshotdate = datetime.datetime.today().strftime('%m_%d_%Y')
root_cache = config['root']['cache']
ksads_cache_path = os.path.join(root_cache, 'ksads')
if not os.path.exists(ksads_cache_path):
    os.mkdir(ksads_cache_path)

root_store = config['root']['store']
# this will be the place to save any snapshots on the nrg servers
store_space = os.path.join(root_store, 'ksads')
if not os.path.exists(store_space):
    os.mkdir(store_space)

# connect to Box
box = LifespanBox(cache=ksads_cache_path, config_file=config['config_files']['box'])

# snapshot folder (used to be the combined folder)
ksads_snapshotfolderid = 48203202724
snapshotQCfolder = 76434619813

# download one of the identical key files which contain the labels for
# all the numbered questions in KSADS
cachekeyfile = box.downloadFile(506958838440)

# hard coding to prevent read of files that shouldnt be in these folders in box.  cant do a search.
# WU,UMN,UCLA, and Harvard, respectively, for cleanest start file ids below
assessments = config['Assessments']


def main():
    for item in assessments:
        # Download latest clean (box) files for each site - make sure they exist.  send up warning if they dont -
        # Before running this program, Montly update process is to download data (3 files - intro,screener,supplement,
        # for each site) from KSADS.net, append rows to site/assessment file (by hand until this program written to automate),
        # and incorporate and errors from wiki...look for any missing rows, but check redcap visit summary to make sure not already known to be missing
        # then save updated xls files with new date, navigate to file in box and select 'upload new version.'
        # This will allow BOX to track versioning until better system is in place
        site_files = assessments[item]['cleanest_start']

        # Get all rows from all site output files for cleanest files as a
        # pandas dataframe with column labels
        rows = get_all_rows(site_files)

        # create snapshot of combined file (all four sites together) store
        # snapshot in 'store' and in 'snapshots' under all sites directory in
        # box.
        snap = 'KSADS_' + item + '_Snapshot_' + snapshotdate + '.csv'
        snapshot_filepath = os.path.join(store_space, snap)  # or in store_space?
        QC_filepath = os.path.join(ksads_cache_path, 'QC_' + snap)
        dictcsv_filepath = os.path.join(ksads_cache_path, 'Dict_' + snap)

        # write rows to csv in store
        rows.to_csv(snapshot_filepath, index=False)
        # upload the snapshot into box
        box.upload_file(snapshot_filepath, ksads_snapshotfolderid)

        # compare ids from snapshot (currently loaded into 'rows' dataframe)
        # with those in Redcap.
        studyids = redcap.getredcapids()
        studydata = redcap.getredcapdata()
        rowsofinterest = rows[['ID', 'PatientID',
                               'PatientType', 'SiteName']].copy()
        new = rowsofinterest['PatientID'].str.split("_", 1, expand=True)
        rowsofinterest['subject'] = new[0].str.strip()
        combowredcap = pd.merge(
            rowsofinterest,
            studyids,
            how='left',
            on='subject')
        # these are the ids that need to be checked by the sites
        combonotinredcap = combowredcap.loc[combowredcap.Subject_ID.isnull()].copy()
        combonotinredcap['reason'] = 'PatientID not in Redcap'
        # combonotinredcap[['ID','PatientID','PatientType','SiteName','reason']].to_csv(QC_filepath,index=False)

        combowredcap2 = pd.merge(
            rowsofinterest,
            studydata,
            how='right',
            on='subject')
        combonotinbox = combowredcap2.loc[combowredcap2.ID.isnull()]
        notinboxunique = combonotinbox.drop_duplicates('subject_id')
        # notinboxunique.loc[(notinboxunique.interview_date<'2019-05-01') & (notinboxunique.flagged.isnull()==True)]
        notinboxunique = notinboxunique.loc[notinboxunique.flagged.isnull()]
        notinboxunique = notinboxunique.loc[notinboxunique.interview_date < '2019-05-01']
        notinboxunique = notinboxunique.loc[~(
            notinboxunique.study == 'hcpa')].copy()
        notinboxunique['reason'] = 'Missing in Box'

        dups = combowredcap.loc[combowredcap.duplicated(
            subset=['PatientID', 'PatientType'], keep=False)]
        dups['reason'] = 'Duplicated ID'

        catQC = pd.concat([combonotinredcap, notinboxunique, dups], axis=0, sort=True)[
            ['ID', 'PatientID', 'subject', 'study', 'site', 'gender', 'interview_date', 'reason']]
        catQC = catQC.sort_values(['site', 'study'])
        catQC.to_csv(QC_filepath, index=False)

        # upload QC file to box
        box.upload_file(QC_filepath, snapshotQCfolder)

        # just keep these in mind for time being
        # these are the ids that need to be
        comboflaggedinredcap = combowredcap.loc[combowredcap.flagged.isnull(
        ) == False]
        # excluded before any data is shared...make note.

        # NDA release will be based on an explicitly named snapshot...not just whatever is in the rows right now - but the following files will
        # ease the process of harmonization.
        # slimhandle = assessments[item]['slim_id'] #this is a temporary file in the Behavioral Data - Snapshots / KSADS / temp folder for use viewing data
        # slimf = makeslim(snapshot_filepath,slimhandle)  #make slim file in cache from snapshot in store (which is identical copy with same named file on box).
        # Note this will overwrite whatever is in the slim file with current data - write slim file to cache and to Scratch Export to NDA
        # make a draft datadictionary from slim file and start comparing columns - write dictionary to cache and to Scratch Export to NDA
        # makedatadictv2(slimf,assessments[item]['dict_id'],cachekeyfile,assessments[item]['key_sheet'])
        inst = item
        makedatadictv2(snapshot_filepath, dictcsv_filepath, inst)
        box.upload_file(dictcsv_filepath, ksads_snapshotfolderid)
    # Clean up cache space
    shutil.rmtree(box.cache)


# screen=finddups(527416597176) #screener snapshot
# duplicatedScreener = screen.loc[screen.duplicated(subset=['PatientID','PatientType'],keep=False)]
# suppl=finddups(527424900084)
# duplicatedSuppl=suppl.loc[suppl.duplicated(subset=['PatientID','PatientType'],keep=False)]


def findupdates(base_id=454918321952, compare_id=454298717674):
    """
    compare two files by dataset id for updates to other columns
    """
    fbase = box.getFileById(base_id)
    basecachefile = box.downloadFile(base_id)
    wb_base = load_workbook(filename=basecachefile)
    basequestionnaire = wb_base[wb_base.sheetnames[0]]
    fbaseraw = pd.DataFrame(basequestionnaire.values)
    header = fbaseraw.iloc[0]
    fbaseraw = fbaseraw[1:]
    fbaseraw.columns = header
    # now the file to compare
    fcompare = box.getFileById(compare_id)
    comparecachefile = box.downloadFile(compare_id)
    wb_compare = load_workbook(filename=comparecachefile)
    comparequestionnaire = wb_compare[wb_compare.sheetnames[0]]
    fcompareraw = pd.DataFrame(comparequestionnaire.values)
    header = fcompareraw.iloc[0]
    fcompareraw = fcompareraw[1:]
    fcompareraw.columns = header
    fjoined = pd.merge(fbaseraw, fcompareraw, on='ID', how='inner')
    # for all columns except the ID, compare...
    updates = pd.DataFrame()
    for col in fbaseraw.columns:
        if col == 'ID':
            pass
        else:
            fjoined.loc[fjoined[str(col) + '_x'] is None] = ""
            fjoined.loc[fjoined[str(col) + '_y'] is None] = ""
            update = fjoined.loc[~(
                fjoined[str(col) + '_x'] == fjoined[str(col) + '_y'])].copy()
            if update.empty:
                pass
            else:
                update['column_affected'] = str(col)
                update['base_value'] = fjoined[str(col) + '_x']
                update['compare_value'] = fjoined[str(col) + '_y']
                updates = updates.append(update)
    updates = updates[['ID', 'PatientID_x', 'SiteName_x',
                       'column_affected', 'base_value', 'compare_value']].copy()
    updates.rename(
        columns={
            'PatientID_x': 'PatientID_base',
            'SiteName_x': 'SiteName_base'})
    updates['basename'] = fbase.get().name
    updates['comparename'] = fcompare.get().name
    updates['base_id'] = base_id
    updates['compare_id'] = compare_id
    return updates


def get_all_rows(sites):
    """
    reads all rows from site xls files as data frame
    """
    dataframes = []
    for site_file in sites:
        # Download file contents to cache
        path = box.downloadFile(site_file)
        df = pd.read_excel(path)
        df.columns = df.columns.astype(str)
        dataframes.append(df)

    return pd.concat(dataframes)


def append_new_data(rows, combined_file):
    """
    Add rows for ids that don't exist and upload to Box
    """
    # print(rows)
    print(str(len(rows)) +
          ' rows found in old or new box files (may contain duplicates)')
    # combined_file_name = combined_file.get().name
    combined_file_path = os.path.join(box.cache, combined_file.get().name)
    # Get existing ids
    existing_ids = []
    with open(combined_file_path) as f:
        for combinedrow in f.readlines():
            # print(row.split(',')[0])
            existing_ids.append(str(combinedrow.split(',')[0]))
    # Get new rows based on id
    new_rows = []
    for row in rows:
        # print('record id: ' + str(row[0]))
        if str(row[0]) not in existing_ids:
            new_rows.append(row)
            # print(combined_file_name)
            # print('record id: ' + str(row[0]))
    print(str(len(new_rows)) + ' new rows')
    if not new_rows:
        print('Nothing new to add. Exiting Append Method...')
        return
    # Write new rows to combined file
    with open(combined_file_path, 'a') as csvfile:
        writer = csv.writer(
            csvfile,
            delimiter=',',
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL
        )
        for row in new_rows:
            writer.writerow(row)
    # Put the file back in Box as a new version
    # box.update_file(combined_file_id, combined_file_name)
    combined_file.update_contents(combined_file_path)


def makeslim(storefilename, slim_id):
    """
    remove columns from cachecopy that have no data and upload slim file to box
    """
    box.downloadFile(slim_id)
    ksadsraw = pd.read_csv(storefilename, header=0, low_memory=False)
    ksadsraw = ksadsraw.dropna(axis=1, how='all')
    snapname = os.path.basename(storefilename)
    combined_fileout = os.path.join(
        box.cache, snapname.split('.')[0] + 'Slim.csv')
    ksadsraw.to_csv(combined_fileout, index=False)
    return box.update_file(slim_id, combined_fileout)


def makedatadictv2(filecsv, dictcsv, inst):
    """
    create datadictionary from csvfile
    """
    ksadsraw = pd.read_csv(filecsv, header=0, low_memory=False)
    varvalues = pd.DataFrame(
        columns=[
            'variable',
            'values_or_example',
            'numunique'])
    varvalues['variable'] = ksadsraw.columns
    kcounts = ksadsraw.count().reset_index().rename(
        columns={'index': 'variable', 0: 'num_nonmissing'})
    varvalues = pd.merge(varvalues, kcounts, on='variable', how='inner')
    summarystats = ksadsraw.describe().transpose(
    ).reset_index().rename(columns={'index': 'variable'})
    varvalues = pd.merge(varvalues, summarystats, on='variable', how='left')
    varvalues['min'] = varvalues['min'].fillna(-99)
    varvalues['max'] = varvalues['max'].fillna(-99)
    varvalues['ValueRange'] = varvalues['min'].astype(int).astype(
        str) + ' :: ' + varvalues['max'].astype(int).astype(str)
    varvalues['min'] = varvalues['min'].replace(-99.0, np.nan)
    varvalues['max'] = varvalues['max'].replace(-99.0, np.nan)
    varvalues.loc[varvalues.ValueRange.str.contains('-99'), 'ValueRange'] = ' '
    # create a data frame containing summary info of data in the ksadraw, e.g.
    # variablse, their formats, values, ect.
    for var in ksadsraw.columns:
        row = ksadsraw.groupby(var).count().reset_index()[var]
        varvalues.loc[varvalues.variable == var, 'numunique'] = len(
            row)  # number of unique vars in this column
        varvalues.loc[(varvalues.variable == var) & (varvalues.numunique <= 10) &
                      (varvalues.num_nonmissing >= 10), 'values_or_example'] = ''.join(
            str(ksadsraw[var].unique().tolist()))
        varvalues.loc[(varvalues.variable == var) & (varvalues.numunique <= 10) & (
            varvalues.num_nonmissing < 10), 'values_or_example'] = ksadsraw[var].unique().tolist()[0]
        varvalues.loc[(varvalues.variable == var) & (varvalues.numunique > 10),
                      'values_or_example'] = ksadsraw[var].unique().tolist()[0]
    varvalues['Instrument Title'] = inst
    varvalues.to_csv(dictcsv, index=False)


def makedatadict(slimf, dict_id, cachekeyfile, sheet):
    """
    create datadictionary from csvfile and upload dictionary to box
    """
    try:
        dictf = box.getFileById(dict_id)
        box.downloadFile(dict_id)
    except BaseException:
        dictf = None
    cachefile = os.path.join(box.cache, slimf.get().name.split('.')[0])
    ksadsraw = pd.read_csv(cachefile + '.csv', header=0, low_memory=False)
    varvalues = pd.DataFrame(
        columns=[
            'variable',
            'values_or_example',
            'numunique'])
    varvalues['variable'] = ksadsraw.columns
    kcounts = ksadsraw.count().reset_index().rename(
        columns={'index': 'variable', 0: 'num_nonmissing'})
    varvalues = pd.merge(varvalues, kcounts, on='variable', how='inner')
    # create a data frame containing summary info of data in the ksadraw, e.g.
    # variablse, their formats, values, ect.
    for var in ksadsraw.columns:
        row = ksadsraw.groupby(var).count().reset_index()[var]
        varvalues.loc[varvalues.variable == var, 'numunique'] = len(
            row)  # number of unique vars in this column
        varvalues.loc[(varvalues.variable == var) & (varvalues.numunique <= 10) &
                      (varvalues.num_nonmissing >= 10), 'values_or_example'] = ''.join(
            str(ksadsraw[var].unique().tolist()))
        varvalues.loc[(varvalues.variable == var) & (varvalues.numunique <= 10) & (
            varvalues.num_nonmissing < 10), 'values_or_example'] = ksadsraw[var].unique().tolist()[1]
        varvalues.loc[(varvalues.variable == var) & (varvalues.numunique > 10),
                      'values_or_example'] = ksadsraw[var].unique().tolist()[1]
    # capture labels for the vars in this assessment from the key
    keyasrow = pd.read_excel(cachekeyfile, sheet_name=sheet, header=0)
    varlabels = keyasrow.transpose().reset_index().rename(
        columns={'index': 'variable', 0: 'question_label'})
    varlabels['variable'] = varlabels['variable'].apply(str)
    # now merge labels for the informative variables from cache
    varvalues2 = pd.merge(varvalues, varlabels, on='variable', how='left')
    varvalues2 = varvalues2[['variable',
                             'question_label',
                             'values_or_example',
                             'numunique',
                             'num_nonmissing']].copy()
    # push this back to box
    fileoutdict = os.path.join(ksads_cache_path, cachefile + "_DataDictionary.csv")
    varvalues2.to_csv(fileoutdict, index=False)
    if dictf is None:
        box.upload_file(fileoutdict, str(ksadscombinedfolderid))
    else:
        box.update_file(dict_id, fileoutdict)


if __name__ == '__main__':
    main()


def share_ksads(
        snapshot=snapshotdate,
        specialstring='_20190813_LS_datarequest',
        boxoutdir=84470235182):
    """
    put in date of snapshot you're useing to create export...e.g. 08_15_2019  #remember all snapshots are copied to store space so no need to grab from BOX
    """
    for item in assessments:
        # create snapshot of combined file (all four sites together) store
        # snapshot in 'store' and in 'snapshots' under all sites directory in
        # box.
        snap = 'KSADS_' + item + '_Snapshot_' + snapshotdate + '.csv'
        snapshort = 'KSADS_' + item + '_Snapshot_' + snapshotdate
        snapshotfile = os.path.join(store_space, snap)
        rows = pd.read_csv(snapshotfile, header=0, low_memory=False)
        studyids = redcap.getredcapids()
        new = rows['PatientID'].str.split("_", 1, expand=True)
        rows['subject'] = new[0].str.strip()
        combowredcap = pd.merge(rows, studyids, how='inner', on='subject')
        combowredcap = combowredcap.loc[combowredcap.PatientID.str.contains(
            'V1')].copy()
        # these are the ones that need to be excluded
        combonotflaggedinredcap = combowredcap.loc[combowredcap.flagged.isnull(
        )]
        temporaryfile = os.path.join(
            ksads_cache_path, snapshort + specialstring + '.csv')
        combonotflaggedinredcap.to_csv(temporaryfile, index=False)
        # upload QC file to box
        box.upload_file(temporaryfile, boxoutdir)

    # Clean up cache space
    shutil.rmtree(box.cache)
