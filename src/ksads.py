import os
import csv
import sys
import shutil
from openpyxl import load_workbook

from download.box import LifespanBox

verbose = True

root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
processed_file = os.path.join(root_dir, 'store/processed-ksads.txt')
cache_space = os.path.join(root_dir, 'cache', 'ksads')
# behavioral_folder_id = 0
behavioral_folder_id = 18445162758
box = LifespanBox(cache=cache_space)

assessments = {
    'Screener': {
        'pattern': '*KSADS*Screener*.xlsx',
        'combined_file_id': 286668148773
    },
    'Intro': {
        'pattern': '*KSADS*Intro*.xlsx',
        'combined_file_id': 286671795350
    },
    'Supplement': {
        'pattern': '*KSADS*Supplement*.xlsx',
        'combined_file_id': 286667260248
    }
}


def main():
    for item in assessments:
        # Download output for each site
        pattern = assessments[item]['pattern']
        results = box.search(pattern=pattern, exclude='Key,MRH')

        # Get all rows from all site output files
        rows = get_all_rows(results)

        # Get combined file
        file_id = assessments[item]['combined_file_id']
        f = box.download_file(file_id)
        # print(f.get().name)

        # Append rows that don't already exist and upload to Box
        append_new_data(rows, f)

    # Clean up cache space
    # shutil.rmtree(box.cache)


def get_all_rows(sites):
    rows = []

    for site_file in sites:
        # Download file contents to cache
        box.download_file(site_file.id)

        path = os.path.join(box.cache, site_file.name)
        wb = load_workbook(filename=path)
        # print(wb.sheetnames)

        if len(wb.sheetnames) > 1:
            print('More than one worksheet.')
            print(wb.sheetnames)
            print('Using the first one --> ' + wb.sheetnames[0])
            continue

        questionnaire = wb[wb.sheetnames[0]]
        print(questionnaire)
        # print(questionnaire.values[0])
        # print(dir(questionnaire.values[0]))
        # sys.exit()

        # Skip the header for all but the first file
        current_row = 0
        for row in questionnaire.values:
            # print(row)
            if current_row != 0:
                rows.append(row)
            current_row += 1

    return rows


def append_new_data(rows, combined_file):
    """
    Add rows for ids that don't exist and upload to Box
    """
    # print(rows)
    print(str(len(rows)) + ' existing rows')
    # combined_file_name = combined_file.get().name
    combined_file_path = os.path.join(box.cache, combined_file.get().name)
    # combined_file_id = f.get().id
    # print(filename)

    # Get existing ids
    existing_ids = []
    with open(combined_file_path) as f:
        for row in f.readlines():
            # print(row.split(',')[0])
            existing_ids.append(str(row.split(',')[0]))
    # print(existing_ids)

    # Get new rows based on id
    new_rows = []
    for row in rows:
        # print('record id: ' + str(row[0]))
        if str(row[0]) not in existing_ids:
            new_rows.append(row)
            # print(combined_file_name)
            # print('record id: ' + str(row[0]))
    print(str(len(new_rows)) + ' new rows')

    if not new_rows:
        print('Nothing new to add. Exiting...')
        return

    # Write new rows to combined file
    with open(combined_file_path, 'a') as csvfile:
        writer = csv.writer(
            csvfile,
            delimiter=',',
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL
        )
        for row in new_rows:
            writer.writerow(row)

    # Put the file back in Box as a new version
    # box.update_file(combined_file_id, combined_file_name)
    combined_file.update_contents(combined_file_path)


if __name__ == '__main__':
    main()
