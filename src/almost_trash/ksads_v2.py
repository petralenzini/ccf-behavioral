import os, datetime
import csv
import pycurl, io
import sys
import shutil
from openpyxl import load_workbook
import pandas as pd
from download.box import LifespanBox

###this stuff needs to be put int a configuration file for defining a redcap object (like a box object)
###...work in progress
study=['hcdparent','hcdchild','hcd18','hca']
token=['CCF4AEC4EFB874DAD5ED47FE8F3B13BB','9C1B8F67F9B83DD073EA5146B5BCC473','16247BDF6EFC2B9D1110F6EB939BAFEE','51660051D09C5F2D923839C099C02903']
field=['parent_id','subject_id','subject_id','subject_id']
event=['visit_1_arm_1','visit_1_arm_1','visit_arm_1','visit_1_arm_1']
###


verbose = True
#verbose = False
snapshotdate = datetime.datetime.today().strftime('%m_%d_%Y')

#root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
root_cache='/data/intradb/tmp/box2nda_cache/'
processed_file = '/home/shared/HCP/hcpinternal/ccf-nda-behavioral/store/processed-ksads_'+snapshotdate+'.txt'
cache_space = os.path.join(root_cache, 'ksads')
os.mkdir(cache_space)

# behavioral_folder_id = 0
behavioral_folder_id = '18445162758'
box = LifespanBox(cache=cache_space)
#study_folders=[18445153825,18445193428,18445162792,18445196396]

#ksadscombinedfolderid=48203202724  #this is the Scratch KSADS folder under Export to NDA in box.
#


#download one of the indentical key files which contain the labels for all the numbered questions in KSADS
keyfileid=box.search(pattern='*Key')
box.download_file(keyfileid[0].id)
qkey=keyfileid[0].name
cachekeyfile=os.path.join(cache_space,qkey)

#hard coding to prevent read of files that shouldnt be in these folders in box.  cant do a search.
#WU,UMN,UCLA, and Harvard, respectively, for cleanest start file ids below
assessments = {
    'Screener': {
        'pattern': '*KSADS*Screener',
#        'combined_file_id': 286668148773,
        'slim_id':449798152073,
        'dict_id':450226557492,
        'key_sheet':'Screener',
        'cleanest_start':[317204091034,317230106899,317224489061,317216055920]  
    },
    'Intro': {
        'pattern': '*KSADS*Intro',
#        'combined_file_id': 286671795350,
        'slim_id':449745327689,
        'dict_id':450219939565,
        'key_sheet':'intro',
        'cleanest_start':[317203291349,317226607001,317224169848,317223373321]
    },
    'Supplement': {
        'pattern': '*KSADS*Supplement',
#        'combined_file_id': 286667260248,
        'slim_id':449779011618,
        'dict_id':450238563945,
        'key_sheet':'Supplement',
        'cleanest_start':[317203382109,317226471328,317224494224,317222199440]
    }
}

def main():
    for item in assessments:
        # Download latest files for each site - make sure they exist.  send up warning if they dont
        site_files=assessments[item]['cleanest_start']

        # Get all rows from all site output files for cleanest files
        rows = get_all_rows(site_files) 

        # create combined snapshot file
        file_id = assessments[item]['combined_file_id']
        f = box.download_file(file_id)
        # print(f.get().name)

        # Append rows to the downloaded combined file that don't already exist and upload back to Box
        append_new_data(rows, f)

         #compare ids from downloaded combined file, with those in Redcap.
parents=getids(token=token[0],field=field[0],event=event[0])
kids=getids(token=token[1],field=field[1],event=event[1])
lateteens=getids(token=token[2],field=field[2],event=event[2])
hca=getids(token=token[3],field=field[3],event=event[3])
currentredcapids=pd.concat([parents,kids,lateteens,hca],axis=0)
        
combined_file_path = os.path.join(box.cache, f.get().name)
combo_before=pd.read_csv(combined_file_path,header=0,low_memory=False)
combo_before=combo_before[['ID','PatientID','PatientType','subject','SiteName']].copy()

new=combo_before['PatientID'].str.split("_",1,expand=True)
combo_before['subject']=new[0].str.strip()
combowredcap=pd.merge(combo_before, currentredcapids, how='left',on='subject')
combonotinredcap=combowredcap.loc[combowredcap.Subject_ID.isnull()==True]
comboflaggedinredcap=combowredcap.loc[combowredcap.flagged.isnull()==False]
        #create list and remove exclusions 

        #now for NDA: remove columns that dont have any data and upload slim file to box
        slimhandle = assessments[item]['slim_id']
        slimf = makeslim(f,slimhandle)
    
        #make a draft datadictionary from slim file and start comparing columns
        makedatadict(slimf,assessments[item]['dict_id'],cachekeyfile,assessments[item]['key_sheet'])

    # Clean up cache space
    shutil.rmtree(box.cache)


def getids(token=token[0],field=field[0],event=event[0]):
    data = {
        'token': token,
        'content': 'record',
        'format': 'csv',
        'type': 'flat',
        'fields[0]': field,
        'events[0]': event,
        'rawOrLabel': 'raw',
        'rawOrLabelHeaders': 'raw',
        'exportCheckboxLabel': 'false',
        'exportSurveyFields': 'false',
        'exportDataAccessGroups': 'false',
        'returnFormat': 'json'
        }
    buf = io.BytesIO()
    ch = pycurl.Curl()
    ch.setopt(ch.URL, 'https://redcap.wustl.edu/redcap/srvrs/prod_v3_1_0_001/redcap/api/')
    ch.setopt(ch.HTTPPOST, list(data.items()))
    ch.setopt(ch.WRITEDATA, buf)
    ch.perform()
    ch.close()
    htmlString = buf.getvalue().decode('UTF-8')
    buf.close()
    parent_ids=pd.DataFrame(htmlString.splitlines(),columns=['Subject_ID'])
    parent_ids=parent_ids.iloc[1:,]
    parent_ids=parent_ids.loc[~(parent_ids.Subject_ID=='')]
    uniqueids=pd.DataFrame(parent_ids.Subject_ID.unique(),columns=['Subject_ID'])
    uniqueids['Subject_ID']=uniqueids.Subject_ID.str.strip('\'"')
    new=uniqueids['Subject_ID'].str.split("_",1,expand=True)
    uniqueids['subject']=new[0].str.strip()
    uniqueids['flagged']=new[1].str.strip()
    return uniqueids


def findupdates(base_id=454918321952,compare_id=454298717674):
    """
    compare two files by dataset id for updates to other columns
    """
    fbase = box.download_file(base_id)
    basecachefile=os.path.join(box.cache,fbase.get().name)
    wb_base = load_workbook(filename=basecachefile)
    basequestionnaire = wb_base[wb_base.sheetnames[0]] 
    fbaseraw=pd.DataFrame(basequestionnaire.values)
    header=fbaseraw.iloc[0]
    fbaseraw=fbaseraw[1:]
    fbaseraw.columns=header
    # now the file to compare
    fcompare = box.download_file(compare_id)
    comparecachefile=os.path.join(box.cache,fcompare.get().name)
    wb_compare = load_workbook(filename=comparecachefile)
    comparequestionnaire = wb_compare[wb_compare.sheetnames[0]] 
    fcompareraw=pd.DataFrame(comparequestionnaire.values)
    header=fcompareraw.iloc[0]
    fcompareraw=fcompareraw[1:]
    fcompareraw.columns=header
    fjoined=pd.merge(fbaseraw,fcompareraw,on='ID',how='inner') 
    #for all columns except the ID, compare...
    updates=pd.DataFrame()
    for col in fbaseraw.columns:
        if col=='ID':
            pass
        else:
            fjoined.loc[fjoined[str(col)+'_x']==None]=""
            fjoined.loc[fjoined[str(col)+'_y']==None]=""
            update=fjoined.loc[~(fjoined[str(col)+'_x']==fjoined[str(col)+'_y'])].copy()
            if update.empty:
                pass
            else:
                update['column_affected']=str(col)
                update['base_value']=fjoined[str(col)+'_x']
                update['compare_value']=fjoined[str(col)+'_y']
                updates=updates.append(update)
    updates=updates[['ID','PatientID_x','SiteName_x','column_affected','base_value','compare_value']].copy()
    updates.rename(columns={'PatientID_x':'PatientID_base','SiteName_x':'SiteName_base'})
    updates['basename']=fbase.get().name
    updates['comparename']=fcompare.get().name
    updates['base_id']=base_id
    updates['compare_id']=compare_id
    return updates



def get_all_rows(sites):
    rows = []
    for site_file in sites:
        # Download file contents to cache
        fh=box.download_file(site_file)#.id)
        path = os.path.join(box.cache, fh.get().name)
        wb = load_workbook(filename=path)
        # print(wb.sheetnames)
        if len(wb.sheetnames) > 1:
            print('More than one worksheet.')
            print(wb.sheetnames)
            print('Using the first one --> ' + wb.sheetnames[0])
            continue
        questionnaire = wb[wb.sheetnames[0]]
        print(questionnaire)
        # Skip the header for all but the first file
        current_row = 0
        for row in questionnaire.values:
            # print(row)
            if current_row != 0:
                rows.append(row)
            current_row += 1
    return rows


def append_new_data(rows, combined_file):
    """
    Add rows for ids that don't exist and upload to Box
    """
    # print(rows)
    print(str(len(rows)) + ' rows found in old or new box files (may contain duplicates)')
    # combined_file_name = combined_file.get().name
    combined_file_path = os.path.join(box.cache, combined_file.get().name)
    # Get existing ids
    existing_ids = []
    with open(combined_file_path) as f:
        for combinedrow in f.readlines():
            # print(row.split(',')[0])
            existing_ids.append(str(combinedrow.split(',')[0]))
    # Get new rows based on id
    new_rows = []
    for row in rows:
        # print('record id: ' + str(row[0]))
        if str(row[0]) not in existing_ids:
            new_rows.append(row)
            # print(combined_file_name)
            # print('record id: ' + str(row[0]))
    print(str(len(new_rows)) + ' new rows')
    if not new_rows:
        print('Nothing new to add. Exiting Append Method...')
        return
    # Write new rows to combined file
    with open(combined_file_path, 'a') as csvfile:
        writer = csv.writer(
            csvfile,
            delimiter=',',
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL
        )
        for row in new_rows:
            writer.writerow(row)
    # Put the file back in Box as a new version
    # box.update_file(combined_file_id, combined_file_name)
    combined_file.update_contents(combined_file_path)


def makeslim(f,slim_id):
    """
    remove columns from cachecopy that have no data and upload slim file to box
    """
    slimf = box.download_file(slim_id)

    cachefile=os.path.join(box.cache,f.get().name)
    ksadsraw=pd.read_csv(cachefile,header=0,low_memory=False)
    ksadsraw=ksadsraw.dropna(axis=1, how='all')
    combined_fileout =  os.path.join(box.cache,f.get().name.split('.')[0]+'Slim.csv')
    ksadsraw.to_csv(combined_fileout,index=False)
   
    #box.client.folder(str(ksadscombinedfolderid)).upload(fileout)
    slimf.update_contents(combined_fileout)

    return slimf

def makedatadict(slimf,dict_id,cachekeyfile,sheet):    
    """
    create datadictionary from csvfile and upload dictionary to box
    """
    try: dictf=box.download_file(dict_id)
    except: dictf=None

    cachefile=os.path.join(box.cache,slimf.get().name.split('.')[0])
    ksadsraw=pd.read_csv(cachefile+'.csv',header=0,low_memory=False)

    varvalues=pd.DataFrame(columns=['variable','values_or_example','numunique'])
    varvalues['variable']=ksadsraw.columns
    kcounts=ksadsraw.count().reset_index().rename(columns={'index':'variable',0:'num_nonmissing'})
    varvalues=pd.merge(varvalues,kcounts,on='variable',how='inner')

    #create a data frame containing summary info of data in the ksadraw, e.g. variablse, their formats, values, ect.
    for var in ksadsraw.columns:
        row=ksadsraw.groupby(var).count().reset_index()[var]
        varvalues.loc[varvalues.variable==var,'numunique']=len(row) #number of unique vars in this column
        varvalues.loc[(varvalues.variable==var) & (varvalues.numunique<=10) & 
            (varvalues.num_nonmissing>=10),'values_or_example']=''.join(str(ksadsraw[var].unique().tolist()))
        varvalues.loc[(varvalues.variable==var) & (varvalues.numunique<=10) & 
            (varvalues.num_nonmissing<10),'values_or_example']=ksadsraw[var].unique().tolist()[1]
        varvalues.loc[(varvalues.variable==var) & (varvalues.numunique>10),'values_or_example']=ksadsraw[var].unique().tolist()[1]

    #capture labels for the vars in this assessment from the key
    keyasrow=pd.read_excel(cachekeyfile,sheet_name=sheet,header=0)
    varlabels=keyasrow.transpose().reset_index().rename(columns={'index':'variable',0:'question_label'})
    varlabels['variable']=varlabels['variable'].apply(str)


    #now merge labels for the informative variables from cache
    varvalues2=pd.merge(varvalues,varlabels,on='variable',how='left')
    varvalues2=varvalues2[['variable','question_label','values_or_example','numunique','num_nonmissing']].copy()
    #push this back to box

    fileoutdict=os.path.join(cache_space,cachefile+"_DataDictionary.csv")
    varvalues2.to_csv(fileoutdict,index=False)

    if dictf==None:
        box.client.folder(str(ksadscombinedfolderid)).upload(fileoutdict)
    else:
        dictf.update_contents(fileoutdict)



if __name__ == '__main__':
    main()
